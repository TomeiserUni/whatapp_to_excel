import io
import json
import os
import re
import secrets
import sys
import tempfile
import threading
import time
import unicodedata
import webbrowser
from collections import Counter, OrderedDict
from datetime import datetime, timezone
from functools import wraps
from pathlib import Path

from flask import Flask, jsonify, redirect, render_template, request, send_file, session

# ── Ambiente ─────────────────────────────────────────────────────
IS_CLOUD  = bool(os.environ.get("RAILWAY_ENVIRONMENT") or os.environ.get("RENDER"))
IS_FROZEN = getattr(sys, "frozen", False)

# ── Caminhos ─────────────────────────────────────────────────────
if IS_FROZEN:
    _BUNDLE   = Path(sys._MEIPASS)
    _USER_DIR = Path.home() / "WhatsAppExcel"
else:
    _BUNDLE   = Path(__file__).resolve().parent.parent
    _USER_DIR = _BUNDLE

sys.path.insert(0, str(_BUNDLE / "src"))

# Na cloud o output é em memória; localmente usa pasta
OUTPUT_DIR = None if IS_CLOUD else _USER_DIR / "output"
if OUTPUT_DIR:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# Registo de linhas não-reconhecidas, para perceber a tendência dos erros
# e priorizar que aliases criar a seguir.
_LOG_NAO_RECONHECIDAS = _USER_DIR / "data" / "nao_reconhecidas.jsonl"

# Aliases aprendidos pelo utilizador (botão "Corrigir" na UI). Ficheiro
# separado do aliases.json curado à mão; é fundido por cima ao carregar.
_ALIASES_APRENDIDOS = _USER_DIR / "data" / "aliases_aprendidos.json"

# Aliases criados deliberadamente na janela de Aliases (aba "Adicionados").
# Permanentes e geridos pela utilizadora; têm prioridade sobre os aprendidos.
_ALIASES_CRIADOS = _USER_DIR / "data" / "aliases_criados.json"

# Palavras de unidade que o cliente escreve a seguir à quantidade ("12 unidades",
# "6 un", "3 pcs"). 'unid\w*' tolera os typos frequentes do WhatsApp ('unidases',
# 'unidaees', 'unidde'…) — qualquer palavra começada por 'unid' conta como unidade.
_UNIDADES_RE = r"(?:unid\w*|un|und|pcs?|cada)"


# Cache em memória dos dois ficheiros de aliases. _linha_so_contexto consulta-os
# por cada linha de cada encomenda; reler o disco de cada vez seria lento. A cache
# é invalidada (posta a None) sempre que se grava.
_cache_aprendidos: dict | None = None
_cache_criados: dict | None = None


def _carregar_aliases_aprendidos() -> dict:
    global _cache_aprendidos
    if _cache_aprendidos is None:
        try:
            with open(_ALIASES_APRENDIDOS, encoding="utf-8") as f:
                _cache_aprendidos = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            _cache_aprendidos = {}
    return _cache_aprendidos


def _gravar_alias_aprendido(chave: str, produto: str) -> None:
    global _cache_aprendidos
    aprendidos = dict(_carregar_aliases_aprendidos())
    aprendidos[chave] = produto
    _ALIASES_APRENDIDOS.parent.mkdir(parents=True, exist_ok=True)
    with open(_ALIASES_APRENDIDOS, "w", encoding="utf-8") as f:
        json.dump(aprendidos, f, ensure_ascii=False, indent=2)
    _cache_aprendidos = None  # invalida


def _carregar_aliases_criados() -> dict:
    global _cache_criados
    if _cache_criados is None:
        try:
            with open(_ALIASES_CRIADOS, encoding="utf-8") as f:
                _cache_criados = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            _cache_criados = {}
    return _cache_criados


def _guardar_aliases_criados(criados: dict) -> None:
    global _cache_criados
    _ALIASES_CRIADOS.parent.mkdir(parents=True, exist_ok=True)
    with open(_ALIASES_CRIADOS, "w", encoding="utf-8") as f:
        json.dump(criados, f, ensure_ascii=False, indent=2)
    _cache_criados = None  # invalida


def _alias_produtos(valor) -> list[str]:
    """
    Normaliza o valor de um alias para uma lista de produtos. Aceita o formato
    antigo (string = 1 produto) e o novo (lista = vários). Marcadores vazios
    ("" / __INEXISTENTE__) ficam de fora — são tratados como "ignorar linha".
    """
    if isinstance(valor, list):
        return [p for p in valor if p and p != INEXISTENTE]
    if isinstance(valor, str) and valor and valor != INEXISTENTE:
        return [valor]
    return []


def _normalizar_para_tendencia(linha: str) -> str:
    """
    Linha sem quantidade/pontuação, minúsculas — agrupa variações do mesmo erro
    e dá chaves de alias estáveis (independentes da quantidade da encomenda).

    Remove números E palavras de unidade em QUALQUER posição: "super mãe 12
    unidades", "12 super mãe" e "super mãe" colapsam todas em "super mae".
    Caso contrário um alias aprendido ficava colado à quantidade daquele pedido
    e só voltava a disparar com exatamente o mesmo número.
    """
    s = re.sub(r"[^\w\s]", " ", linha.lower())              # pontuação → espaço
    s = re.sub(rf"\b{_UNIDADES_RE}\b", " ", s)              # palavras de unidade
    s = re.sub(r"\b\d+\b", " ", s)                          # quantidades (qq posição)
    return re.sub(r"\s+", " ", s).strip()


def _registar_nao_reconhecida(linha: str) -> None:
    """Acrescenta uma linha não-reconhecida ao log (JSONL). Falha em silêncio."""
    chave = _normalizar_para_tendencia(linha)
    if not chave or _linha_so_contexto(linha):
        return
    try:
        _LOG_NAO_RECONHECIDAS.parent.mkdir(parents=True, exist_ok=True)
        registo = {
            "ts": datetime.now(timezone.utc).isoformat(timespec="seconds"),
            "linha": linha.strip(),
            "chave": chave,
        }
        with open(_LOG_NAO_RECONHECIDAS, "a", encoding="utf-8") as f:
            f.write(json.dumps(registo, ensure_ascii=False) + "\n")
    except Exception:
        pass  # registo é best-effort, nunca deve quebrar o processamento

# ── Flask ─────────────────────────────────────────────────────────
template_folder = str(_BUNDLE / "src" / "templates")
app = Flask(__name__, template_folder=template_folder)
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50 MB
app.secret_key = os.environ.get("SECRET_KEY") or secrets.token_hex(32)

_pipeline = None


def _int_env(name: str, default: int, minimum: int = 1, maximum: int | None = None) -> int:
    try:
        value = int(os.environ.get(name, default))
    except (TypeError, ValueError):
        value = default
    value = max(minimum, value)
    if maximum is not None:
        value = min(value, maximum)
    return value


def _chunks(items: list[str], size: int) -> list[list[str]]:
    return [items[i:i + size] for i in range(0, len(items), size)]


def _normalizar_token(texto: str) -> str:
    texto = "".join(
        c for c in unicodedata.normalize("NFD", texto.lower())
        if unicodedata.category(c) != "Mn"
    )
    return texto.strip()


_GENERICOS_SO_CONTEXTO = {
    "cores novas", "verniz normal", "verniz gel", "verniz gel cateye",
}


def _tem_alias_para_linha(linha: str) -> bool:
    """
    True se a linha corresponde a um alias criado/aprendido ativo. Um alias que a
    utilizadora criou de propósito (ex: "cores novas" → 6 vernizes) tem prioridade
    sobre a lista de linhas-contexto: deixa de ser ruído e passa a ser um pedido.
    """
    chave = _normalizar_para_tendencia(linha)
    if not chave:
        return False
    if chave in _carregar_aliases_criados():
        return True
    # Aprendidos só contam se apontarem para um produto real (não __INEXISTENTE__).
    valor = _carregar_aliases_aprendidos().get(chave)
    return bool(_alias_produtos(valor))


def _linha_so_contexto(linha: str) -> bool:
    lower = linha.lower().strip()
    lower_sem_qtd = re.sub(r"\b\d+\b", "", lower)
    lower_sem_qtd = re.sub(rf"\b{_UNIDADES_RE}\b", "", lower_sem_qtd)
    # remove pontuação/hífens soltos (ex: "4- verniz gel cateye" → "verniz gel cateye")
    lower_sem_qtd = re.sub(r"[^\w\s]", " ", lower_sem_qtd)
    lower_sem_qtd = re.sub(r"\s+", " ", lower_sem_qtd).strip()
    # Um alias criado/aprendido para esta linha anula a regra de contexto.
    if _tem_alias_para_linha(linha):
        return False
    return (
        lower in {"de cada"} or lower_sem_qtd in _GENERICOS_SO_CONTEXTO
        or lower.startswith(("bom dia", "boa tarde", "boa noite"))
        or lower.startswith(("encomenda ", "pedido "))
        or "cores novas" in lower
    )


def _linha_origem(linhas: list[str], trecho: str) -> str:
    trecho = (trecho or "").strip()
    if not trecho:
        return ""
    trecho_lower = trecho.lower()
    for idx, linha in enumerate(linhas, 1):
        if trecho_lower == linha.lower() or trecho_lower in linha.lower():
            return f"linha {idx}: {linha}"
    return trecho


def _linha_idx(linhas: list[str], trecho: str) -> int | None:
    trecho = (trecho or "").strip()
    if not trecho:
        return None
    trecho_lower = trecho.lower()
    for idx, linha in enumerate(linhas, 1):
        if trecho_lower == linha.lower() or trecho_lower in linha.lower():
            return idx
    return None


# Marcador especial: a colega confirmou via botão Corrigir que este texto
# não corresponde a nenhum produto da loja. Gravado como alias para "".
INEXISTENTE = "__INEXISTENTE__"


def _erro_linha(idx: int, linha: str) -> dict:
    chave = _normalizar_para_tendencia(linha)
    inexistente = _carregar_aliases_aprendidos().get(chave) == INEXISTENTE
    return {
        "ficheiro":     "texto colado",
        "produto":      "",
        # quantidade deduzida do texto, para o caso de a colega corrigir a linha
        "qtd":          _quantidade_linha(linha),
        "score":        0,
        "ref":          "",
        "texto_origem": f"linha {idx}: {linha}",
        "inexistente":  inexistente,
    }


def _intercalar_erros_linhas(linhas: list[str], resultados: list[dict]) -> list[dict]:
    por_linha: dict[int, list[dict]] = {}
    sem_linha: list[dict] = []

    for row in resultados:
        idx = row.pop("_linha_idx", None)
        if idx is None:
            sem_linha.append(row)
            continue
        por_linha.setdefault(idx, []).append(row)

    final = []
    for idx, linha in enumerate(linhas, 1):
        rows = por_linha.get(idx)
        if rows:
            final.extend(rows)
        else:
            final.append(_erro_linha(idx, linha))
            _registar_nao_reconhecida(linha)
    final.extend(sem_linha)
    return final


def _dedupe_resultados_por_linha(resultados: list[dict]) -> list[dict]:
    por_linha: dict[int, dict] = {}
    sem_linha: list[dict] = []

    for row in resultados:
        idx = row.get("_linha_idx")
        if idx is None:
            sem_linha.append(row)
            continue
        atual = por_linha.get(idx)
        if atual is None or float(row.get("score", 0)) > float(atual.get("score", 0)):
            por_linha[idx] = row

    return list(por_linha.values()) + sem_linha


def _dedupe_resultados_por_linha_ou_de_cada(resultados: list[dict], linhas: list[str]) -> list[dict]:
    por_linha: dict[int, dict[str, dict]] = {}
    sem_linha: list[dict] = []

    for row in resultados:
        idx = row.get("_linha_idx")
        if idx is None:
            sem_linha.append(row)
            continue

        produtos_linha = por_linha.setdefault(idx, {})
        produto = row.get("produto", "")
        atual = produtos_linha.get(produto)
        if atual is None or float(row.get("score", 0)) > float(atual.get("score", 0)):
            produtos_linha[produto] = row

    final = []
    for produtos_linha in por_linha.values():
        final.extend(produtos_linha.values())
    final.extend(sem_linha)
    return final


def _quantidade_de_cada(linha: str) -> int:
    antes = linha.lower().split("de cada", 1)[0]
    nums = re.findall(r"\b\d+\b", antes)
    return int(nums[-1]) if nums else 1


def _quantidade_linha(linha: str, ignorar: set[str] | None = None) -> int:
    """
    Deduz a quantidade da linha. `ignorar` é um conjunto de números que NÃO
    devem ser tratados como quantidade — tipicamente identificadores de produto
    já reconhecidos (ex: em "6 like gel 139, 197, 120" o 139/197/120 são produtos,
    a quantidade é o 6).
    """
    ignorar = ignorar or set()
    m = re.search(rf"\b(\d+)\s*{_UNIDADES_RE}\b", linha, re.IGNORECASE)
    if m:
        return int(m.group(1))
    m = re.search(r"\bx\s*(\d+)\b|\b(\d+)\s*x\b", linha, re.IGNORECASE)
    if m:
        return int(m.group(1) or m.group(2))
    nums = [
        m.group(1)
        for m in re.finditer(r"\b(\d+)\b", linha)
        if m.group(1) not in ignorar
        and not re.match(r"\s*(?:g|gr|gramas?|ml)\b", linha[m.end():], re.IGNORECASE)
    ]
    if not nums:
        return 1
    # Vários números "soltos" (sem unidade) e nenhum é quantidade explícita:
    # a quantidade costuma ser o PRIMEIRO (ex: "6 like gel 139, 197, 120" → 6),
    # não o último, que seria mais um identificador.
    return int(nums[0])


def _token_e_quantidade(linha: str, match: re.Match) -> bool:
    after = linha[match.end():match.end() + 12]
    before = linha[max(0, match.start() - 2):match.start()]
    return bool(
        re.match(rf"\s*{_UNIDADES_RE}\b", after, re.IGNORECASE)
        or re.search(r"x\s*$", before, re.IGNORECASE)
    )


def _singular(token: str) -> str:
    """Normaliza plurais portugueses simples para o singular, para que
    'recargas'≡'recarga', 'brocas'≡'broca', 'lixas'≡'lixa'. Mantém números."""
    if token.isdigit() or len(token) < 4:
        return token
    if token.endswith("oes"):     # botoes→botao
        return token[:-3] + "ao"
    if token.endswith("aes"):     # paes→pao
        return token[:-3] + "ao"
    if token.endswith("ais"):     # casais→casal
        return token[:-3] + "al"
    if token.endswith("eis"):     # papeis→papel
        return token[:-3] + "el"
    if token.endswith("ns"):      # garrafons→garrafom (raro)
        return token[:-2] + "m"
    if token.endswith("s"):       # recargas→recarga, brocas→broca
        return token[:-1]
    return token


def _mapa_tokens_unicos(produtos: list[str]) -> dict[str, str]:
    token_produto: dict[str, str] = {}
    repetidos: set[str] = set()

    for produto in produtos:
        for token in {_singular(t) for t in re.findall(r"\w+", _normalizar_token(produto))}:
            if len(token) < 3 and not token.isdigit():
                continue
            if token in token_produto and token_produto[token] != produto:
                repetidos.add(token)
                continue
            token_produto[token] = produto

    return {token: produto for token, produto in token_produto.items() if token not in repetidos}


def _base_comum(nomes) -> str:
    """Prefixo de palavras comum a vários nomes, parando no 1º número/diferença.
    Ex: ["like gel 139 roxo", "like gel 197 chili"] → "like gel"."""
    listas = [n.lower().split() for n in nomes]
    if not listas:
        return ""
    base = []
    for tokens in zip(*listas):
        t = tokens[0]
        if t.isdigit() or any(x != t for x in tokens):
            break
        base.append(t)
    return " ".join(base)


def _expandir_identificadores_unicos(linhas: list[str], produtos: list[str], sku_map: dict,
                                     ignorar_linhas: set[int] | None = None) -> list[dict]:
    unicos = _mapa_tokens_unicos(produtos)
    ignorar_linhas = ignorar_linhas or set()
    resultados = []

    for idx, linha in enumerate(linhas, 1):
        if idx in ignorar_linhas or _linha_so_contexto(linha):
            continue

        produtos_linha: dict[str, str] = {}
        linha_norm = _normalizar_token(linha)
        for match in re.finditer(r"\w+", linha_norm):
            token = match.group(0)
            if token.isdigit() and _token_e_quantidade(linha_norm, match):
                continue
            produto = unicos.get(_singular(token))
            if produto:
                produtos_linha[produto] = token

        if not produtos_linha:
            continue

        # Desambiguação por base: se algum produto foi reconhecido por número
        # (ex: "like gel 104" → base "like gel"), usar essa base para resolver os
        # outros números da linha que ficaram por reconhecer por serem ambíguos
        # sozinhos (ex: "120" em "like gel 104, 120" → "like gel 120 ...").
        produtos_por_num = {tok: prod for prod, tok in produtos_linha.items() if tok.isdigit()}
        if produtos_por_num:
            base = _base_comum(produtos_por_num.values())
            if base:
                for match in re.finditer(r"\b(\d+)\b", linha_norm):
                    num = match.group(1)
                    if num in produtos_por_num or _token_e_quantidade(linha_norm, match):
                        continue
                    candidatos = [
                        p for p in produtos
                        if num in p.lower().split() and p.lower().startswith(base)
                    ]
                    if len(candidatos) == 1:
                        produtos_linha[candidatos[0]] = num

        # Números que são identificadores de produto não contam como quantidade.
        nums_produto = {tok for tok in produtos_linha.values() if tok.isdigit()}
        qty = _quantidade_linha(linha, ignorar=nums_produto)
        for produto in produtos_linha:
            resultados.append({
                "ficheiro":     "texto colado",
                "produto":      produto,
                "qtd":          qty,
                "score":        1.0,
                "ref":          sku_map.get(produto, ""),
                "texto_origem": f"linha {idx}: {linha}",
                "_linha_idx":   idx,
            })

    return resultados


def _expandir_aliases_diretos(linhas: list[str], aliases: dict, sku_map: dict) -> list[dict]:
    resultados = []
    # Cada alias pode mapear para 1 produto (string) ou vários (lista). Guarda-se
    # já como lista de produtos; aliases sem produtos úteis (ex: marcador "") saem.
    aliases_ordenados = sorted(
        ((alias, _alias_produtos(valor)) for alias, valor in (aliases or {}).items()),
        key=lambda item: len(item[0]),
        reverse=True,
    )
    aliases_ordenados = [(a, prods) for a, prods in aliases_ordenados if prods]

    for idx, linha in enumerate(linhas, 1):
        if _linha_so_contexto(linha):
            continue

        linha_norm = _normalizar_token(linha)
        for alias, produtos_alias in aliases_ordenados:
            alias_norm = _normalizar_token(alias)
            if not alias_norm:
                continue
            m_alias = re.search(rf"(?<!\w){re.escape(alias_norm)}(?!\w)", linha_norm)
            if not m_alias:
                continue

            # Aliases multi-palavra genéricos (ex: "rosa leitoso") não devem
            # disparar quando são só um pedaço de um produto maior na linha
            # (ex: "fiber base rosa leitoso cintilante" → deve ir à AI). Só
            # aplica se as palavras da linha fora do alias forem insignificantes.
            palavras_alias = alias_norm.split()
            if len(palavras_alias) >= 2:
                resto = (linha_norm[:m_alias.start()] + " " + linha_norm[m_alias.end():])
                extra = [w for w in resto.split() if not w.isdigit() and len(w) >= 3]
                if len(extra) >= 2:
                    continue  # linha tem contexto a mais → deixa a AI decidir

            # Quantidade do produto. Uma qty com unidade EXPLÍCITA na linha
            # ("12 unidades", "6 un") é a quantidade deste produto, esteja onde
            # estiver — o alias pode vir antes dela com palavras pelo meio
            # ("gloss top coat 12 unidades" → 12; "super mãe 12 unidades" → 12).
            # Sem unidade explícita, restringe-se ao segmento até ao fim do alias
            # (a qty costuma vir antes: "2 like gel 104"); números soltos depois
            # pertencem a outros produtos da linha.
            nums_alias = set(re.findall(r"\d+", alias_norm))
            m_qty_unidade = re.search(
                rf"\b(\d+)\s*{_UNIDADES_RE}\b", linha_norm, flags=re.IGNORECASE,
            )
            if m_qty_unidade and m_qty_unidade.group(1) not in nums_alias:
                qtd = int(m_qty_unidade.group(1))
            else:
                qtd = _quantidade_linha(linha_norm[:m_alias.end()], ignorar=nums_alias)
            # Um alias pode expandir para vários produtos (ex: "recargas drill" →
            # 3 lixas); todos recebem a mesma quantidade da linha.
            for produto in produtos_alias:
                resultados.append({
                    "ficheiro":     "texto colado",
                    "produto":      produto,
                    "qtd":          qtd,
                    "score":        1.0,
                    "ref":          sku_map.get(produto, ""),
                    "texto_origem": f"linha {idx}: {linha}",
                    "_linha_idx":   idx,
                })
            break

    return resultados


_RE_LISTA_VARIANTES = re.compile(
    r"""^\s*(?:\d+\s+)?                              # qty inicial opcional
        ([a-záàâãéèêíóôõúç ]+?)\s+                    # base (palavras alfabéticas)
        (\d+(?:\s*,\s*\d+){1,})\s+                    # 2+ números separados por vírgula
        (\d+)\s*                                       # qty final
        (?:un\w*|pcs?|cada|de\s+cada)\b""",
    re.IGNORECASE | re.VERBOSE,
)


def _expandir_lista_variantes_numeradas(linhas: list[str], produtos: list[str], sku_map: dict) -> list[dict]:
    """
    Padrão "BASE N1,N2,N3,... QTY [un|cada|de cada]" → expande para BASE+N1, BASE+N2, ...
    Cada variante recebe a quantidade QTY.

    Ex: "Acrigel 5,6,7,9 6 unidades cada" →
        polyacrygel 5/6/7/9 com qty=6 cada (4 produtos).

    Para cada número, escolhe o único produto do catálogo que tem esse número como token
    e cuja base faz match (fuzz ≥ 70). Se houver ambiguidade, salta esse número.
    """
    from rapidfuzz import fuzz

    resultados = []
    for idx, linha in enumerate(linhas, 1):
        m = _RE_LISTA_VARIANTES.match(linha)
        if not m:
            continue
        base = m.group(1).strip().lower()
        nums = re.findall(r"\d+", m.group(2))
        qty = int(m.group(3))
        if len(nums) < 2 or not base:
            continue

        for num in nums:
            candidatos = []
            for p in produtos:
                if num not in p.lower().split():
                    continue
                score = max(
                    fuzz.partial_ratio(base, p.lower()),
                    fuzz.token_set_ratio(base, p.lower()),
                )
                if score >= 70:
                    candidatos.append((p, score))
            candidatos.sort(key=lambda x: -x[1])
            # único candidato OU primeiro com margem clara sobre o segundo
            if not candidatos:
                continue
            if len(candidatos) > 1 and candidatos[0][1] - candidatos[1][1] < 10:
                continue
            produto = candidatos[0][0]
            resultados.append({
                "ficheiro":     "texto colado",
                "produto":      produto,
                "qtd":          qty,
                "score":        1.0,
                "ref":          sku_map.get(produto, ""),
                "texto_origem": f"linha {idx}: {linha}",
                "_linha_idx":   idx,
            })
    return resultados


def _expandir_regras_de_cada(linhas: list[str], sku_map: dict) -> list[dict]:
    produtos_moldes_f1 = [
        "moldes f1 quadrado",
        "silicones para moldes f1",
        "moldes f1 amendoa russa",
    ]

    resultados = []
    for idx, linha in enumerate(linhas, 1):
        lower = linha.lower()
        if "moldes f1" not in lower or "de cada" not in lower:
            continue

        qty = _quantidade_de_cada(linha)
        for produto in produtos_moldes_f1:
            resultados.append({
                "ficheiro":     "texto colado",
                "produto":      produto,
                "qtd":          qty,
                "score":        1.0,
                "ref":          sku_map.get(produto, ""),
                "texto_origem": f"linha {idx}: {linha}",
                "_linha_idx":   idx,
            })
    return resultados


# Código SKU: NN.NN.NNN com pontos, espaços ou traços como separador.
_RE_SKU = re.compile(r"(?<!\d)(\d{2})[.\s-](\d{2})[.\s-](\d{3})(?!\d)")


def _expandir_referencias_sku(linhas: list[str], sku_map: dict) -> list[dict]:
    """
    Linha que contém um código SKU (ex: "3 - 93.01.018") → produto desse SKU.
    A quantidade vem do número antes do código. SKU inexistente → não dispara.
    """
    ref_para_produto = {ref: nome for nome, ref in sku_map.items() if ref}
    resultados = []
    for idx, linha in enumerate(linhas, 1):
        for m in _RE_SKU.finditer(linha):
            ref = f"{m.group(1)}.{m.group(2)}.{m.group(3)}"
            produto = ref_para_produto.get(ref)
            if not produto:
                continue
            antes = linha[:m.start()]
            m_qty = re.search(r"(\d+)\s*[-.)]?\s*$", antes)
            qty = int(m_qty.group(1)) if m_qty else 1
            resultados.append({
                "ficheiro":     "texto colado",
                "produto":      produto,
                "qtd":          qty,
                "score":        1.0,
                "ref":          ref,
                "texto_origem": f"linha {idx}: {linha}",
                "_linha_idx":   idx,
            })
    return resultados


def _mapa_brocas_numeradas(produtos: list[str]) -> dict[str, str]:
    """{ '12': '12 - broca flecha ...' } para os produtos broca numerados (N - broca …)."""
    mapa: dict[str, str] = {}
    for p in produtos:
        m = re.match(r"^(\d+)\s*-\s*broca\b", p.lower())
        if m:
            mapa[m.group(1)] = p
    return mapa


def _expandir_brocas_numeradas(linhas: list[str], produtos: list[str], sku_map: dict) -> list[dict]:
    """
    "broca(s) N" ou "N broca(s)" → produto numerado "N - broca …".
    Ex: "3 brocas 12" → '12 - broca flecha …' com qty=3.
    O número que casa a broca é o identificador; a quantidade vem do outro número.
    """
    mapa = _mapa_brocas_numeradas(produtos)
    if not mapa:
        return []

    resultados = []
    for idx, linha in enumerate(linhas, 1):
        lower = linha.lower()
        m_broca = re.search(r"\bbrocas?\b", lower)
        if not m_broca:
            continue
        # O número ANTES de "broca(s)" é quantidade; os DEPOIS são identificadores
        # (ex: "3 brocas 12" → qty 3, broca 12). Evita tratar a qty como broca nº.
        ids_broca = [
            m.group(1) for m in re.finditer(r"\b(\d+)\b", lower)
            if m.start() > m_broca.start() and m.group(1) in mapa
        ]
        if not ids_broca:
            continue
        m_qty = re.search(r"(\d+)\s*$", lower[:m_broca.start()])
        qty = int(m_qty.group(1)) if m_qty else 1
        for num in dict.fromkeys(ids_broca):  # únicos, mantém ordem
            produto = mapa[num]
            resultados.append({
                "ficheiro":     "texto colado",
                "produto":      produto,
                "qtd":          qty,
                "score":        1.0,
                "ref":          sku_map.get(produto, ""),
                "texto_origem": f"linha {idx}: {linha}",
                "_linha_idx":   idx,
            })
    return resultados


def _melhor_linha_origem(produto: str, linhas: list[str], fuzz_mod, todos_produtos: list[str] = None) -> str:
    genericas = {
        "verniz", "gel", "normal", "unidade", "unidades", "un", "pcs", "cada",
        "base", "top", "coat", "ml", "g", "gr",
    }
    if todos_produtos:
        token_counts: dict[str, int] = {}
        for p in todos_produtos:
            for t in re.findall(r"\w+", p.lower()):
                if len(t) >= 3 and not t.isdigit():
                    token_counts[t] = token_counts.get(t, 0) + 1
        genericas = genericas | {t for t, c in token_counts.items() if c > 1}
    produto_lower = produto.lower()
    produto_tokens = set(re.findall(r"\w+", produto_lower))
    distintivas = {t for t in produto_tokens if t not in genericas and not t.isdigit()}
    numeros_produto = {t for t in produto_tokens if t.isdigit()}

    melhor_linha = ""
    melhor_score = -1.0
    for linha in linhas:
        if _linha_so_contexto(linha):
            continue
        linha_lower = linha.lower()
        linha_tokens = set(re.findall(r"\w+", linha_lower))
        score = fuzz_mod.partial_ratio(produto_lower, linha_lower)
        score += fuzz_mod.token_set_ratio(produto_lower, linha_lower) * 0.5
        score += len(distintivas & linha_tokens) * 80
        score += len(numeros_produto & linha_tokens) * 120
        if score > melhor_score:
            melhor_linha = linha
            melhor_score = score

    return melhor_linha


def _enable_ai_pipeline(api_key: str) -> None:
    if _pipeline is None:
        return

    import anthropic
    import ai_pipeline as ai_pl

    client = anthropic.Anthropic(api_key=api_key)
    ai_prods, ai_sku, ai_aliases, exemplos = ai_pl.load_produtos(_BUNDLE / "data")
    _pipeline.update({
        "ai_pl":      ai_pl,
        "ai_client":  client,
        "ai_produtos": ai_prods,
        "ai_sku_map": ai_sku,
        "ai_aliases": ai_aliases,
        "exemplos":   exemplos,
    })


# ── Autenticação ──────────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if os.environ.get("APP_PASSWORD") and not session.get("authenticated"):
            return redirect("/login")
        return f(*args, **kwargs)
    return decorated


@app.route("/login", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        username = request.form.get("username", "")
        password = request.form.get("password", "")
        ok_user = (username == os.environ.get("APP_USERNAME", ""))
        ok_pass = (password == os.environ.get("APP_PASSWORD", ""))
        if ok_user and ok_pass:
            session["authenticated"] = True
            return redirect("/")
        error = "Utilizador ou password incorretos."
    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    session.clear()
    return redirect("/login")


def _aplicar_aliases_aprendidos() -> None:
    """
    Funde os aliases criados (janela Aliases) e aprendidos (botão Corrigir) por
    cima dos curados, em memória. Prioridade: criados > aprendidos > curados.
    """
    if _pipeline is None:
        return
    aprendidos = _carregar_aliases_aprendidos()
    criados = _carregar_aliases_criados()
    if not aprendidos and not criados:
        return
    # Marcadores de "não existe na loja" → alias para "" (linha ignorada no matching).
    def _efetivo(v):
        return "" if v == INEXISTENTE else v
    efetivos = {k: _efetivo(v) for k, v in aprendidos.items()}
    efetivos.update({k: _efetivo(v) for k, v in criados.items()})  # criados ganham
    for campo in ("aliases", "ai_aliases"):
        base = _pipeline.get(campo)
        if isinstance(base, dict):
            _pipeline[campo] = {**base, **efetivos}


def _carregar_catalogo_loja() -> tuple[list, dict] | None:
    """
    Catálogo completo da Shopkit em RAM (uma chamada no arranque). É a fonte de
    produtos quando há SHOPKIT_API_KEY — substitui o prod.pkl. Devolve (nomes,
    sku_map) ou None se não houver key. Salvaguarda em disco gerida pela própria
    shopkit_api (usada se a loja estiver indisponível).
    """
    shop_key = os.environ.get("SHOPKIT_API_KEY")
    if not shop_key:
        return None
    import shopkit_api
    return shopkit_api.carregar_catalogo(shop_key)


def _load_pipeline():
    global _pipeline
    from dotenv import load_dotenv
    load_dotenv()

    api_key = os.environ.get("ANTHROPIC_API_KEY")

    # Fonte do catálogo: a Shopkit (uma chamada no arranque) quando há key; o
    # catálogo local (.pkl/ai_pipeline.load_produtos) só entra se a loja não estiver
    # configurada. A pasta data/ deixa de ser fonte que se mantém à mão.
    catalogo_loja = _carregar_catalogo_loja()

    if IS_CLOUD:
        import anthropic
        import ai_pipeline as ai_pl
        client = anthropic.Anthropic(api_key=api_key)
        prod_local, sku_local, aliases, exemplos = ai_pl.load_produtos(_BUNDLE / "data")
        produtos, sku_map = catalogo_loja if catalogo_loja else (prod_local, sku_local)
        _pipeline = {"pl": ai_pl, "ai_pl": ai_pl, "produtos": produtos, "emb_prod": None,
                     "sku_map": sku_map, "aliases": aliases, "exemplos": exemplos, "ai_client": client}
    elif catalogo_loja or not (_BUNDLE / "data" / "prod.pkl").exists():
        # Sem pipeline ML local: o catálogo vem da loja e o ai_pipeline trata texto e
        # imagem (rapidfuzz + regras + Claude sobre o catálogo em RAM). Sem embeddings.
        if not api_key:
            raise RuntimeError(
                "Este modo requer ANTHROPIC_API_KEY definida — sem Claude não há "
                "como fazer o matching dos produtos."
            )
        import anthropic
        import ai_pipeline as ai_pl
        client = anthropic.Anthropic(api_key=api_key)
        _, sku_local, aliases, exemplos = ai_pl.load_produtos(_BUNDLE / "data")
        if catalogo_loja:
            produtos, sku_map = catalogo_loja
            print(f"[pipeline] Catálogo da loja em RAM: {len(produtos)} produtos (data/ não é fonte).")
        else:
            produtos, sku_map = [], sku_local
            print("[pipeline] Sem catálogo da loja nem prod.pkl: a depender de aliases/SKUs.")
        _pipeline = {"pl": ai_pl, "ai_pl": ai_pl, "produtos": produtos, "emb_prod": None,
                     "sku_map": sku_map, "aliases": aliases, "exemplos": exemplos,
                     "ai_client": client,
                     "ai_produtos": produtos, "ai_sku_map": sku_map, "ai_aliases": aliases}
    else:
        # Local com pipeline ML para imagens (EasyOCR + embeddings) a partir do .pkl.
        import pipeline as pl
        produtos, emb_prod, sku_map, freq_palavras, palavras_unicas, aliases = pl.load_produtos()
        _pipeline = {"pl": pl, "produtos": produtos, "emb_prod": emb_prod,
                     "sku_map": sku_map, "aliases": aliases, "ai_client": None,
                     "freq_palavras": freq_palavras, "palavras_unicas": palavras_unicas,
                     "ai_pl": None, "exemplos": []}

        # Se tiver API key: carrega também ai_pipeline para texto (muito mais rápido)
        if api_key:
            import anthropic
            import ai_pipeline as ai_pl
            client = anthropic.Anthropic(api_key=api_key)
            ai_prods, ai_sku, ai_aliases, exemplos = ai_pl.load_produtos(_BUNDLE / "data")
            pl.init_ai_client(api_key)
            _pipeline.update({
                "ai_pl":      ai_pl,
                "ai_client":  client,
                "ai_produtos": ai_prods,
                "ai_sku_map": ai_sku,
                "ai_aliases": ai_aliases,
                "exemplos":   exemplos,
            })
            print("[AI] Claude ativo para texto (batches paralelos).")

    _aplicar_aliases_aprendidos()
    print("[pipeline] Pronto.")


# ── Pré-processamento de texto ─────────────────────────────────────
def _expandir_lista_qty_partilhada(linha: str) -> list[str] | None:
    """
    "Prod1, Prod2 e Prod3 50g-16un" → ["Prod1 16", "Prod2 16", "Prod3 16"]
    Só expande se houver pelo menos uma vírgula na parte dos produtos.
    """
    m = re.search(
        r'(?:[\d,.]+\s*(?:g|gr|gramas?|ml)\s*-\s*)?(\d+)\s*(?:un(?:idades?)?|pcs?)\s*$',
        linha, re.IGNORECASE
    )
    if not m:
        return None
    qty = m.group(1)
    parte = linha[:m.start()].strip().rstrip(',').strip()
    if ',' not in parte:
        return None
    nomes = re.split(r'\s*,\s*|\s+e\s+', parte, flags=re.IGNORECASE)
    nomes = [n.strip() for n in nomes if n.strip()]
    if len(nomes) < 2:
        return None
    return [f"{nome} {qty}" for nome in nomes]


def _preprocessar_texto(texto: str) -> list[str]:
    """
    1. Une linhas de continuação (começam com ',' ou 'e ', ou a linha anterior
       termina em ',' ou ' e').
    2. Expande listas com quantidade partilhada no fim.
       Ex: "Gel transparente ,\nbranco leitoso\n,branco leitoso intenso\nPanacota 50g-16un"
           → ["Gel transparente 16", "branco leitoso 16", "branco leitoso intenso 16", "Panacota 16"]
    """
    raw = [l.strip() for l in texto.splitlines()]

    joined = []
    buf = ""
    for l in raw:
        if not l:
            if buf:
                joined.append(buf)
                buf = ""
            continue

        if buf and l.lower() == "de cada":
            buf = f"{buf} de cada"
            continue

        l_sep = l.startswith(',') or l.lower().startswith('e ')
        b_sep = buf.endswith(',') or buf.lower().endswith(' e')

        if buf and (l_sep or b_sep):
            # limpar fim do buf
            b = buf.rstrip()
            if b.endswith(','):
                b = b[:-1].rstrip()
            elif b.lower().endswith(' e'):
                b = b[:-2].rstrip()
            # limpar início de l
            c = l.lstrip()
            if c.startswith(','):
                c = c[1:].lstrip()
            elif c.lower().startswith('e '):
                c = c[2:].lstrip()
            buf = b + ', ' + c
        else:
            if buf:
                joined.append(buf)
            buf = l
    if buf:
        joined.append(buf)

    resultado = []
    for linha in joined:
        exp = _expandir_lista_qty_partilhada(linha)
        resultado.extend(exp if exp else [linha])
    return resultado


# ── Rotas ─────────────────────────────────────────────────────────
@app.route("/")
@login_required
def index():
    return render_template("index.html")


@app.route("/status")
def status():
    # Sem login_required: é o healthcheck do Render e só revela se o pipeline
    # carregou (booleano), não expõe dados sensíveis.
    return jsonify({"ready": _pipeline is not None})


@app.route("/tendencias")
@login_required
def tendencias():
    """Linhas não-reconhecidas mais frequentes — ajuda a priorizar que aliases criar."""
    contagem: Counter = Counter()
    exemplos: dict[str, str] = {}
    total = 0
    try:
        with open(_LOG_NAO_RECONHECIDAS, encoding="utf-8") as f:
            for linha_log in f:
                linha_log = linha_log.strip()
                if not linha_log:
                    continue
                try:
                    reg = json.loads(linha_log)
                except json.JSONDecodeError:
                    continue
                chave = reg.get("chave", "")
                if not chave:
                    continue
                contagem[chave] += 1
                total += 1
                exemplos.setdefault(chave, reg.get("linha", chave))
    except FileNotFoundError:
        pass

    top = [
        {"chave": chave, "vezes": vezes, "exemplo": exemplos.get(chave, chave)}
        for chave, vezes in contagem.most_common(100)
    ]
    return jsonify({"total": total, "distintas": len(contagem), "top": top})


def _catalogo_produtos() -> tuple[list, dict]:
    """Lista de produtos + sku_map ativos no pipeline (qualquer modo)."""
    if _pipeline is None:
        return [], {}
    produtos = _pipeline.get("ai_produtos") or _pipeline.get("produtos") or []
    sku_map = _pipeline.get("ai_sku_map") or _pipeline.get("sku_map") or {}
    return produtos, sku_map


@app.route("/produtos")
@login_required
def produtos():
    """Catálogo local (nome + ref) — estado inicial/fallback do seletor Corrigir."""
    prods, sku_map = _catalogo_produtos()
    return jsonify([{"produto": p, "ref": sku_map.get(p, "")} for p in prods])


def _pesquisar_local(termo: str, limite: int = 15) -> list[dict]:
    prods, sku_map = _catalogo_produtos()
    q = _normalizar_token(termo)
    if not q:
        return [{"produto": p, "ref": sku_map.get(p, "")} for p in prods[:limite]]
    hits = [p for p in prods if q in _normalizar_token(p)]
    return [{"produto": p, "ref": sku_map.get(p, "")} for p in hits[:limite]]


# Cache LRU em memória das pesquisas à API (evita bater na Shopkit repetidamente).
# chave normalizada -> (timestamp, resultados). Teto fixo: ao exceder, descarta-se
# a entrada menos recentemente usada — memória constante, sem crescimento sem fim.
_CACHE_PESQUISA: "OrderedDict[str, tuple[float, list]]" = OrderedDict()
_CACHE_PESQUISA_TTL = 600   # 10 minutos
_CACHE_PESQUISA_MAX = 200   # nº máximo de termos guardados


def _cache_pesquisa_guardar(chave: str, agora: float, resultados: list) -> None:
    _CACHE_PESQUISA[chave] = (agora, resultados)
    _CACHE_PESQUISA.move_to_end(chave)  # marca como usada há menos tempo
    while len(_CACHE_PESQUISA) > _CACHE_PESQUISA_MAX:
        _CACHE_PESQUISA.popitem(last=False)  # descarta a mais antiga (LRU)


def _limpar_quantidade_termo(termo: str) -> str:
    """
    Tira a quantidade do termo para a pesquisa por nome (o filtro AND eliminaria
    todos os resultados se a qty lá estivesse). Remove:
      1) o número INICIAL ("12 super primer" → "super primer");
      2) "número + unidade" em qualquer posição ("super mãe 12 unidades" →
         "super mãe"; "cola 3 un" → "cola").
    Números internos do nome SEM unidade a seguir são preservados ("like gel
    139", "lâmpada 90"), porque aí o número é distintivo, não uma quantidade.
    """
    termo = (termo or "").strip()
    termo = re.sub(r"^\s*\d+\s*[-.)x]?\s*", "", termo, flags=re.IGNORECASE).strip()
    termo = re.sub(rf"\b\d+\s*{_UNIDADES_RE}\b", " ", termo, flags=re.IGNORECASE)
    return re.sub(r"\s+", " ", termo).strip()


def _buscar_candidatos_catalogo(termo: str) -> list[dict] | None:
    """
    Candidatos para uma linha, pesquisados no catálogo JÁ em RAM (mesma filtragem
    AND da Shopkit, mas sem rede). É a fonte usada no processamento de encomendas:
    o catálogo completo foi carregado no arranque, por isso não há razão para bater
    na API por cada linha — o que, em encomendas grandes, gerava 429 (rate limit) e
    fazia linhas falharem ao calhar. Devolve list[{produto,ref}] ou None se não
    houver catálogo em RAM (aí o ai_pipeline cai no rapidfuzz sobre prods_t).
    """
    termo = _limpar_quantidade_termo(termo)
    if not termo:
        return None
    produtos, sku_map = _catalogo_produtos()
    if not produtos:
        return None
    import shopkit_api
    return shopkit_api.pesquisar_em_catalogo(termo, produtos, sku_map)


def _pesquisar_shopkit_cached(termo: str) -> list[dict] | None:
    """
    Pesquisa um termo na Shopkit AO VIVO, com cache LRU. Usada pelo seletor
    'Corrigir' (/pesquisar_produto), onde a colega pode procurar por sinónimos
    que não estão no nome ('clear'→'transparente') e a API os apanha pela descrição.
    Devolve list[{produto, ref}] (possivelmente vazia) ou None em erro/sem API key.
    """
    termo = _limpar_quantidade_termo(termo)
    api_key = os.environ.get("SHOPKIT_API_KEY")
    if not (api_key and termo):
        return None
    chave = _normalizar_token(termo)
    agora = time.time()
    em_cache = _CACHE_PESQUISA.get(chave)
    if em_cache and agora - em_cache[0] < _CACHE_PESQUISA_TTL:
        _CACHE_PESQUISA.move_to_end(chave)
        return em_cache[1]
    try:
        import shopkit_api
        resultados = shopkit_api.pesquisar(api_key, termo)
        _cache_pesquisa_guardar(chave, agora, resultados)
        return resultados
    except Exception as e:
        print(f"[shopkit] pesquisa falhou para {termo!r} ({e}); fallback local.")
        return None


@app.route("/pesquisar_produto")
@login_required
def pesquisar_produto():
    """
    Pesquisa para o seletor 'Corrigir'. Usa a Shopkit como fonte (mesma lógica e
    cache do pipeline, via _pesquisar_shopkit_cached: limpa a quantidade inicial),
    e cai no catálogo local (.pkl) quando a loja não dá candidatos — por erro
    (None) ou por 0 resultados ([]) — para a colega ver sempre opções.
    """
    termo = (request.args.get("q") or "").strip()
    resultados = _pesquisar_shopkit_cached(termo)
    if resultados:
        return jsonify({"fonte": "api", "produtos": resultados})
    return jsonify({"fonte": "local", "produtos": _pesquisar_local(termo)})


@app.route("/aprender_alias", methods=["POST"])
@login_required
def aprender_alias():
    """Recebe {texto, produto}: grava um alias e aplica-o de imediato em memória."""
    dados = request.get_json(silent=True) or {}
    texto = (dados.get("texto") or "").strip()
    produto = (dados.get("produto") or "").strip()
    ref_enviada = (dados.get("ref") or "").strip()
    if not texto or not produto:
        return jsonify({"ok": False, "erro": "texto e produto são obrigatórios"}), 400

    chave = _normalizar_para_tendencia(texto)
    if not chave:
        return jsonify({"ok": False, "erro": "texto sem conteúdo útil"}), 400

    if produto == INEXISTENTE:
        # A colega marcou que este texto não existe na loja.
        _gravar_alias_aprendido(chave, INEXISTENTE)
        _aplicar_aliases_aprendidos()
        return jsonify({"ok": True, "chave": chave, "inexistente": True})

    # Produto pode vir do catálogo local ou de uma pesquisa ao vivo na API.
    # A ref do catálogo tem prioridade; senão usa-se a que veio da pesquisa.
    _, sku_map = _catalogo_produtos()
    ref = sku_map.get(produto) or ref_enviada

    _gravar_alias_aprendido(chave, produto)
    _aplicar_aliases_aprendidos()
    return jsonify({"ok": True, "chave": chave, "produto": produto, "ref": ref})


def _alias_para_ui(valor) -> dict:
    """Formata um alias para a janela: lista de {produto, ref} ou marca inexistente."""
    if valor == INEXISTENTE:
        return {"inexistente": True, "produtos": []}
    _, sku_map = _catalogo_produtos()
    produtos = _alias_produtos(valor)
    return {
        "inexistente": False,
        "produtos": [{"produto": p, "ref": sku_map.get(p, "")} for p in produtos],
    }


@app.route("/aliases", methods=["GET"])
@login_required
def listar_aliases():
    """Aliases para a janela de gestão, separados por aba: criados e aprendidos."""
    criados = {k: _alias_para_ui(v) for k, v in _carregar_aliases_criados().items()}
    aprendidos = {k: _alias_para_ui(v) for k, v in _carregar_aliases_aprendidos().items()}
    return jsonify({"criados": criados, "aprendidos": aprendidos})


@app.route("/aliases", methods=["POST"])
@login_required
def criar_alias():
    """
    Cria/atualiza um alias permanente (aba 'Adicionados').
    Recebe {expressao, produtos: ["nome1", "nome2", ...]}. A chave é normalizada
    como nas tendências (sem quantidade/pontuação), igual ao botão Corrigir.
    """
    dados = request.get_json(silent=True) or {}
    expressao = (dados.get("expressao") or "").strip()
    produtos = [p.strip() for p in (dados.get("produtos") or []) if (p or "").strip()]
    if not expressao or not produtos:
        return jsonify({"ok": False, "erro": "expressão e pelo menos um produto são obrigatórios"}), 400

    chave = _normalizar_para_tendencia(expressao)
    if not chave:
        return jsonify({"ok": False, "erro": "expressão sem conteúdo útil"}), 400

    criados = _carregar_aliases_criados()
    # Guarda string quando é só um produto (retrocompatível), lista quando vários.
    criados[chave] = produtos[0] if len(produtos) == 1 else produtos
    _guardar_aliases_criados(criados)
    _aplicar_aliases_aprendidos()
    return jsonify({"ok": True, "chave": chave, "alias": _alias_para_ui(criados[chave])})


@app.route("/aliases", methods=["DELETE"])
@login_required
def apagar_alias():
    """Apaga um alias. Recebe {chave, aba} onde aba ∈ {'criados','aprendidos'}."""
    dados = request.get_json(silent=True) or {}
    chave = (dados.get("chave") or "").strip()
    aba = (dados.get("aba") or "").strip()
    if not chave or aba not in ("criados", "aprendidos"):
        return jsonify({"ok": False, "erro": "chave e aba válida são obrigatórias"}), 400

    global _cache_aprendidos
    if aba == "criados":
        dados_aba = dict(_carregar_aliases_criados())
        if chave in dados_aba:
            del dados_aba[chave]
            _guardar_aliases_criados(dados_aba)
    else:
        dados_aba = dict(_carregar_aliases_aprendidos())
        if chave in dados_aba:
            del dados_aba[chave]
            _ALIASES_APRENDIDOS.parent.mkdir(parents=True, exist_ok=True)
            with open(_ALIASES_APRENDIDOS, "w", encoding="utf-8") as f:
                json.dump(dados_aba, f, ensure_ascii=False, indent=2)
            _cache_aprendidos = None  # invalida

    # Reconstrói o pipeline a partir do catálogo base + aliases restantes, senão
    # um alias apagado continuava em memória (a fusão só acrescenta, não remove).
    _recarregar_aliases_no_pipeline()
    return jsonify({"ok": True})


def _recarregar_aliases_no_pipeline() -> None:
    """Repõe os aliases curados (do disco) e volta a fundir criados+aprendidos.
    Necessário ao apagar: a fusão em memória só acrescenta, nunca remove."""
    if _pipeline is None:
        return
    try:
        _, _, aliases_curados, _ = _pipeline["ai_pl"].load_produtos(_BUNDLE / "data")
    except Exception as e:
        print(f"[aliases] não consegui recarregar curados: {e}")
        aliases_curados = {}
    for campo in ("aliases", "ai_aliases"):
        if isinstance(_pipeline.get(campo), dict):
            _pipeline[campo] = dict(aliases_curados)
    _aplicar_aliases_aprendidos()


@app.route("/config", methods=["GET"])
@login_required
def get_config():
    key = os.environ.get("ANTHROPIC_API_KEY", "")
    return jsonify({"configured": bool(key)})


@app.route("/config", methods=["POST"])
@login_required
def set_config():
    data = request.json or {}
    env_path = _USER_DIR / ".env"
    from dotenv import set_key as dotenv_set_key

    key = data.get("api_key", "").strip()
    if key:
        dotenv_set_key(str(env_path), "ANTHROPIC_API_KEY", key)
        os.environ["ANTHROPIC_API_KEY"] = key
        if _pipeline is not None and not IS_CLOUD:
            _pipeline["pl"].init_ai_client(key)
            _enable_ai_pipeline(key)

    username = data.get("app_username", "").strip()
    if username:
        dotenv_set_key(str(env_path), "APP_USERNAME", username)
        os.environ["APP_USERNAME"] = username

    password = data.get("app_password", "").strip()
    if password:
        dotenv_set_key(str(env_path), "APP_PASSWORD", password)
        os.environ["APP_PASSWORD"] = password

    return jsonify({"ok": True})


@app.route("/processar", methods=["POST"])
@login_required
def processar():
    if _pipeline is None:
        return jsonify({"error": "Modelos ainda a carregar…"}), 503

    files = request.files.getlist("imagens")
    pl = _pipeline
    resultados = []

    for f in files:
        with tempfile.NamedTemporaryFile(suffix=Path(f.filename).suffix, delete=False) as tmp:
            f.save(tmp.name)
            tmp_path = Path(tmp.name)
        try:
            if IS_CLOUD:
                rows, ocr_text = pl["pl"].processar_imagem(tmp_path, pl["produtos"], pl["sku_map"], pl["aliases"], pl["ai_client"], pl.get("exemplos", []))
                linhas_img = [l for l in _preprocessar_texto(ocr_text or "") if l.strip()]

                rows_img: list[dict] = []
                for row in (rows or []):
                    produto = row[0]
                    score   = row[1]
                    qty     = row[2]
                    ai_idx  = row[3] if len(row) > 3 else None
                    ai_ref  = row[4] if len(row) > 4 else ""
                    if ai_idx and 1 <= ai_idx <= len(linhas_img):
                        linha_idx  = ai_idx
                        texto_orig = f"linha {ai_idx}: {linhas_img[ai_idx - 1]}"
                    else:
                        texto_orig = (ocr_text or "")[:300]
                        linha_idx  = None
                    rows_img.append({
                        "ficheiro":     f.filename,
                        "produto":      produto,
                        "qtd":          qty,
                        "score":        round(score, 3),
                        "ref":          ai_ref or pl["sku_map"].get(produto, ""),
                        "texto_origem": texto_orig,
                        "_linha_idx":   linha_idx,
                    })

                # Regras PLN também sobre o OCR — alinhadas com /processar_texto
                aliases_img = pl.get("aliases", {})
                regras: list[dict] = []
                regras.extend(_expandir_aliases_diretos(linhas_img, aliases_img, pl["sku_map"]))
                regras.extend(_expandir_referencias_sku(linhas_img, pl["sku_map"]))
                brocas_img = _expandir_brocas_numeradas(linhas_img, pl["produtos"], pl["sku_map"])
                regras.extend(brocas_img)
                regras.extend(_expandir_identificadores_unicos(linhas_img, pl["produtos"], pl["sku_map"],
                                                               ignorar_linhas={r["_linha_idx"] for r in brocas_img}))
                regras.extend(_expandir_lista_variantes_numeradas(linhas_img, pl["produtos"], pl["sku_map"]))
                regras.extend(_expandir_regras_de_cada(linhas_img, pl["sku_map"]))
                for r in regras:
                    r["ficheiro"] = f.filename

                combinado = _dedupe_resultados_por_linha_ou_de_cada(rows_img + regras, linhas_img)
                if combinado:
                    resultados.extend(combinado)
                else:
                    resultados.append({"ficheiro": f.filename, "produto": "", "qtd": "", "score": 0, "ref": "", "texto_origem": ""})
            else:
                rows = pl["pl"].processar_imagem(tmp_path, pl["produtos"], pl["emb_prod"],
                                                  pl.get("freq_palavras"), pl.get("palavras_unicas"))
                if rows:
                    for produto, score, qty, trecho in rows:
                        resultados.append({
                            "ficheiro":      f.filename,
                            "produto":       produto,
                            "qtd":           qty,
                            "score":         round(score, 3),
                            "ref":           pl["sku_map"].get(produto, ""),
                            "texto_origem":  trecho or "",
                        })
                else:
                    resultados.append({"ficheiro": f.filename, "produto": "", "qtd": "", "score": 0, "ref": "", "texto_origem": ""})
        finally:
            tmp_path.unlink(missing_ok=True)

    return jsonify(resultados)


@app.route("/processar_texto", methods=["POST"])
@login_required
def processar_texto():
    if _pipeline is None:
        return jsonify({"error": "Modelos ainda a carregar…"}), 503

    texto = (request.json or {}).get("texto", "").strip()
    if not texto:
        return jsonify([])

    pl = _pipeline
    linhas = [l for l in _preprocessar_texto(texto) if l.strip()]
    resultados = []

    ai_pl     = pl.get("ai_pl")
    ai_client = pl.get("ai_client")
    use_ai    = bool(ai_pl and ai_client)

    if use_ai:
        # Estratégia: regras PLN primeiro (determinísticas, score 1.0).
        # AI só processa linhas que nenhuma regra cobriu — mais barato e sem alucinações em linhas já resolvidas.
        from concurrent.futures import ThreadPoolExecutor, as_completed as _as_completed
        from rapidfuzz import fuzz as _rfuzz

        prods_t   = pl["ai_produtos"] if not IS_CLOUD else pl["produtos"]
        sku_t     = pl["ai_sku_map"]  if not IS_CLOUD else pl["sku_map"]
        aliases_t = pl["ai_aliases"]  if not IS_CLOUD else pl["aliases"]
        exemplos_t = pl.get("exemplos", [])
        sku_final = pl.get("ai_sku_map") or pl["sku_map"]

        # 1) Regras PLN sobre todas as linhas
        alias_rows = _expandir_aliases_diretos(linhas, aliases_t, sku_final)
        resultados.extend(alias_rows)
        resultados.extend(_expandir_referencias_sku(linhas, sku_final))
        brocas = _expandir_brocas_numeradas(linhas, prods_t, sku_final)
        resultados.extend(brocas)
        # Linhas já resolvidas por um alias direto não devem ser "enriquecidas" por
        # outras regras (ex: alias "cores novas" → 6 cores; sem isto o identificador
        # único apanhava "nova" e juntava "verniz gel maria nova iorque").
        linhas_alias = {r["_linha_idx"] for r in alias_rows}
        linhas_brocas = {r["_linha_idx"] for r in brocas}
        resultados.extend(_expandir_identificadores_unicos(linhas, prods_t, sku_final, ignorar_linhas=linhas_brocas | linhas_alias))
        resultados.extend(_expandir_lista_variantes_numeradas(linhas, prods_t, sku_final))
        resultados.extend(_expandir_regras_de_cada(linhas, sku_final))

        linhas_cobertas = {r["_linha_idx"] for r in resultados if r.get("_linha_idx")}

        # 2) Linhas que sobraram (e que não são puro contexto) vão à AI
        linhas_para_ai = [
            (idx, linha)
            for idx, linha in enumerate(linhas, 1)
            if idx not in linhas_cobertas and not _linha_so_contexto(linha)
        ]
        print(f"[PLN] {len(linhas_cobertas)}/{len(linhas)} linhas resolvidas por regras. "
              f"{len(linhas_para_ai)} linhas para a AI.")

        melhor: dict[str, tuple] = {}
        if linhas_para_ai:
            batch_size = _int_env("AI_BATCH_LINES", 60, minimum=10, maximum=120)
            max_workers = _int_env("AI_MAX_WORKERS", 8, minimum=1, maximum=12)
            # Cada batch mantém o índice global no formato "N: texto" (lido por _linhas_com_indices)
            linhas_numeradas = [f"{idx}: {linha}" for idx, linha in linhas_para_ai]
            batches = _chunks(linhas_numeradas, batch_size)
            print(f"[AI texto] {len(linhas_para_ai)} linhas em {len(batches)} lote(s) de até {batch_size}.")

            raw_rows: list[tuple] = []
            batch_errors: list[str] = []
            # Fonte dos candidatos: pesquisa no catálogo em RAM (mesma filtragem
            # AND da Shopkit, sem rede). Limpa a quantidade e devolve o nome exato
            # do produto. Sem isto, linhas como "Super mãe 12 unidades" davam
            # candidatos ambíguos no fuzzy (vários produtos com "mãe" empatados a
            # 50) e a AI, sem candidato único, ignorava a linha. Em falha cai no
            # rapidfuzz sobre prods_t (dentro do ai_pipeline).
            buscar = _buscar_candidatos_catalogo
            with ThreadPoolExecutor(max_workers=min(len(batches), max_workers)) as pool:
                futs = {
                    pool.submit(
                        ai_pl.processar_texto,
                        "\n".join(b),
                        prods_t, sku_t, aliases_t, ai_client, exemplos_t,
                        buscar,
                    ): b
                    for b in batches
                }
                for fut in _as_completed(futs):
                    b = futs[fut]
                    try:
                        for row in (fut.result() or []):
                            raw_rows.append((row, b))
                    except Exception as e:
                        print(f"[batch] erro: {e}")
                        batch_errors.append(str(e))

            if batch_errors and len(batch_errors) == len(batches):
                return jsonify({"error": "Erro ao processar com AI. Verifica a API key ou tenta novamente."}), 502

            for row, b in raw_rows:
                p = row[0]
                if p not in melhor or row[1] > melhor[p][0][1]:
                    melhor[p] = (row, b)

        if melhor:
            for p, (row, b) in melhor.items():
                _, s, q = row[0], row[1], row[2]
                ai_linha_idx = row[3] if len(row) > 3 else None
                ai_ref = row[4] if len(row) > 4 else ""
                if ai_linha_idx and 1 <= ai_linha_idx <= len(linhas):
                    linha_idx = ai_linha_idx
                    texto_orig = f"linha {ai_linha_idx}: {linhas[ai_linha_idx - 1]}"
                else:
                    origem = _melhor_linha_origem(p, b, _rfuzz, prods_t)
                    linha_idx = _linha_idx(linhas, origem)
                    texto_orig = _linha_origem(linhas, origem)
                resultados.append({
                    "ficheiro":     "texto colado",
                    "produto":      p,
                    "qtd":          q,
                    "score":        round(s, 3),
                    "ref":          ai_ref or sku_final.get(p, ""),
                    "texto_origem": texto_orig,
                    "_linha_idx":   linha_idx,
                })

        if resultados:
            resultados = _intercalar_erros_linhas(linhas, _dedupe_resultados_por_linha_ou_de_cada(resultados, linhas))
        else:
            return jsonify(_intercalar_erros_linhas(linhas, []))
    else:
        # Sem API key: pipeline local ML (lento para ordens grandes)
        texto_proc = "\n".join(linhas)
        rows = pl["pl"].processar_texto(texto_proc, pl["produtos"], pl["emb_prod"],
                                         pl.get("freq_palavras"), pl.get("palavras_unicas"))
        resultados.extend(_expandir_aliases_diretos(linhas, pl.get("aliases", {}), pl["sku_map"]))
        resultados.extend(_expandir_referencias_sku(linhas, pl["sku_map"]))
        brocas = _expandir_brocas_numeradas(linhas, pl["produtos"], pl["sku_map"])
        resultados.extend(brocas)
        resultados.extend(_expandir_identificadores_unicos(linhas, pl["produtos"], pl["sku_map"],
                                                           ignorar_linhas={r["_linha_idx"] for r in brocas}))
        resultados.extend(_expandir_lista_variantes_numeradas(linhas, pl["produtos"], pl["sku_map"]))
        resultados.extend(_expandir_regras_de_cada(linhas, pl["sku_map"]))
        if rows:
            for row in rows:
                p, s, q = row[0], row[1], row[2]
                trecho = row[3] if len(row) > 3 else ""
                resultados.append({
                    "ficheiro":     "texto colado",
                    "produto":      p,
                    "qtd":          q,
                    "score":        round(s, 3),
                    "ref":          pl["sku_map"].get(p, ""),
                    "texto_origem": _linha_origem(linhas, trecho),
                    "_linha_idx":   _linha_idx(linhas, trecho),
                })
            resultados = _intercalar_erros_linhas(linhas, _dedupe_resultados_por_linha_ou_de_cada(resultados, linhas))
        else:
            return jsonify(_intercalar_erros_linhas(linhas, _dedupe_resultados_por_linha_ou_de_cada(resultados, linhas)))

    return jsonify(resultados)


@app.route("/exportar", methods=["POST"])
@login_required
def exportar():
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment

    FILL_GREEN  = PatternFill("solid", fgColor="C6EFCE")
    FILL_YELLOW = PatternFill("solid", fgColor="FFEB9C")
    FILL_RED    = PatternFill("solid", fgColor="FFC7CE")

    def _fill(score, produto):
        if not produto:
            return FILL_RED
        if score >= 0.90:
            return FILL_GREEN
        if score >= 0.70:
            return FILL_YELLOW
        return FILL_RED

    data = request.json or []
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultados"
    ws.append(["Referência", "Produto", "Quantidade", "Onde foi identificado", "Texto reconhecido"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 34
    ws.column_dimensions["E"].width = 60

    for row in data:
        produto = row.get("produto", "")
        score   = float(row.get("score", 0))
        texto   = row.get("texto_origem", "")
        ficheiro = row.get("ficheiro", "")
        onde = ficheiro
        if texto:
            onde = f"{ficheiro} - {texto}" if ficheiro else texto
        label   = produto if produto else "ERRO — produto não identificado"
        ws.append([row.get("ref", ""), label, row.get("qtd", ""), onde, texto])
        fill = _fill(score, produto)
        for cell in ws[ws.max_row]:
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=False)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="resultados.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ── Main ──────────────────────────────────────────────────────────
if __name__ == "__main__":
    threading.Thread(target=_load_pipeline, daemon=True).start()
    PORT = int(os.environ.get("PORT", 5000 if sys.platform == "win32" else 5002))
    HOST = "0.0.0.0"
    if not IS_CLOUD:
        threading.Timer(1.5, lambda: webbrowser.open(f"http://localhost:{PORT}")).start()
    from waitress import serve
    serve(app, host=HOST, port=PORT)
